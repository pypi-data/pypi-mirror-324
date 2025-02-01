from abc import abstractmethod
import os
import re
from typing import List, Dict, Any, Callable
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from .core import BaseFileGenerator, MultiFileGenerationError
from openpyxl.worksheet.worksheet import Worksheet
import copy


class BaseDocxGenerator(BaseFileGenerator):
    """
    Generates DOCX files from templates using docxtpl.
    Subclasses must implement:
    - get_documents(): Provides template paths and metadata for filenames.
    - _get_output_filename(): Generates the output filename from document info.
    """

    @abstractmethod
    def get_context(self) -> dict:
        pass

    @abstractmethod
    def get_documents(self) -> List[Dict]:
        """
        Returns a list of dictionaries. Each dict must include 'template_path' and
        any additional data needed for generating the output filename.
        """
        pass

    @abstractmethod
    def _get_output_filename(self, doc_info: Dict) -> str:
        """
        Generates the output filename based on document info and instance data.
        """
        pass

    def generate_files_and_return_paths(self) -> List[str]:
        documents = self.get_documents()
        if not documents:
            raise MultiFileGenerationError("No documents found to generate DOCX files.")

        rendered_paths = []
        for doc_info in documents:
            template_path = doc_info.get('template_path')
            if not template_path:
                raise MultiFileGenerationError("Document info missing 'template_path'.")

            full_template_path = self._resolve_template_path(template_path)
            output_filename = self._get_output_filename(doc_info)
            output_path = os.path.join("/tmp", output_filename)
            self._render_template(full_template_path, output_path)
            rendered_paths.append(output_path)

        return rendered_paths

    def _resolve_template_path(self, template_path: str) -> str:
        """Override this method if custom template path resolution is needed."""
        return template_path

    def _render_template(self, template_path: str, output_path: str):
        try:
            template = DocxTemplate(template_path)
            context = self.get_context()
            template.render(context)
            template.save(output_path)
        except Exception as e:
            raise MultiFileGenerationError(f"Error rendering DOCX template '{template_path}': {e}") from e

class BaseXlsxGenerator(BaseFileGenerator):
    """
    Generates XLSX files from templates using openpyxl, supporting context data and loops.
    Ensures:
      • Rows below loop sections remain in place
      • Cell styles are preserved for inserted rows
      • Loop template rows are removed once processed
    """

    @abstractmethod
    def get_context(self) -> Dict[str, Any]:
        """Subclasses must provide context data for template rendering"""
        pass

    @abstractmethod
    def get_template_path(self) -> str:
        """Subclasses must provide absolute path to template file"""
        pass

    @abstractmethod
    def _get_output_filename(self) -> str:
        """Subclasses must define output filename generation logic"""
        pass

    def _resolve_template_path(self, template_path: str) -> str:
        """Can be overridden for custom path resolution logic"""
        return template_path

    def generate_files_and_return_paths(self) -> List[str]:
        context = self.get_context()
        template_path = self.get_template_path()
        full_template_path = self._resolve_template_path(template_path)

        output_filename = self._get_output_filename()
        output_path = os.path.join("/tmp", output_filename)

        try:
            wb = load_workbook(full_template_path)
            ws = wb.active

            self._process_static_placeholders(ws, context)
            self._process_loops(ws, context)

            wb.save(output_path)
            return [output_path]
        except Exception as e:
            raise MultiFileGenerationError(
                f"XLSX generation failed: {e}"
            ) from e

    # All subsequent methods remain unchanged from original implementation
    # Only Django-specific references were removed from the top-level methods

    def _process_static_placeholders(self, worksheet: Worksheet, context: Dict[str, Any]):
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value:
                    replaced_value = self._replace_placeholders(cell.value, context)
                    if replaced_value != cell.value:
                        cell.value = replaced_value

    def _replace_placeholders(self, text: str, context: Dict[str, Any]) -> str:
        return re.sub(
            r"\{\{(\w+)\}\}",
            lambda m: str(context.get(m.group(1), "")),
            text
        )

    def _process_loops(self, worksheet: Worksheet, context: Dict[str, Any]):
        loop_markers = self._find_loop_markers(worksheet)
        for marker_row, marker_col, list_name in sorted(loop_markers, reverse=True):
            self._process_loop(worksheet, marker_row, marker_col, list_name, context)

    def _find_loop_markers(self, worksheet: Worksheet) -> List[tuple]:
        markers = []
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    match = re.match(r"\{\{for\s+\w+\s+in\s+(\w+)\}\}", cell.value.strip())
                    if match:
                        markers.append((cell.row, cell.column, match.group(1)))
        return markers

    def _process_loop(
        self,
        worksheet: Worksheet,
        marker_row: int,
        marker_col: int,
        list_name: str,
        context: Dict[str, Any]
    ):
        items = context.get(list_name, [])
        template_row_idx = marker_row + 1

        if not items:
            worksheet.delete_rows(marker_row, 2)
            return

        template_cells = list(worksheet.iter_rows(
            min_row=template_row_idx,
            max_row=template_row_idx
        ))[0]

        row_style_data = self._extract_row_style_data(worksheet, template_row_idx)
        worksheet.delete_rows(marker_row, 2)
        worksheet.insert_rows(marker_row, len(items))

        for i, item in enumerate(items):
            insert_row_idx = marker_row + i
            self._copy_row_with_style(
                worksheet,
                row_style_data,
                source_row_idx=template_row_idx,
                target_row_idx=insert_row_idx
            )

            row_cells = list(worksheet.iter_rows(min_row=insert_row_idx, max_row=insert_row_idx))[0]
            for src_cell, tgt_cell in zip(template_cells, row_cells):
                if isinstance(src_cell.value, str):
                    replaced = self._replace_item_placeholders(src_cell.value, item)
                    tgt_cell.value = replaced
                else:
                    tgt_cell.value = src_cell.value

    def _replace_item_placeholders(self, text: str, item: Dict[str, Any]) -> str:
        return re.sub(
            r"\{\{item\.(\w+)\}\}",
            lambda m: str(item.get(m.group(1), "")),
            text
        )

    def _extract_row_style_data(self, worksheet: Worksheet, row_idx: int) -> Dict[int, dict]:
        row_style_data = {}
        row_cells = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        row_dim = worksheet.row_dimensions.get(row_idx)
        row_height = row_dim.height if row_dim else None

        for cell in row_cells:
            style_dict = {
                "font": copy.copy(cell.font),
                "border": copy.copy(cell.border),
                "fill": copy.copy(cell.fill),
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection),
                "alignment": copy.copy(cell.alignment),
            }
            row_style_data[cell.column] = style_dict

        row_style_data["_row_height"] = row_height
        return row_style_data

    def _copy_row_with_style(
        self,
        worksheet: Worksheet,
        row_style_data: Dict[int, dict],
        source_row_idx: int,
        target_row_idx: int
    ):
        if row_style_data.get("_row_height") is not None:
            worksheet.row_dimensions[target_row_idx].height = row_style_data["_row_height"]

        target_cells = list(worksheet.iter_rows(min_row=target_row_idx, max_row=target_row_idx))[0]
        for tgt_cell in target_cells:
            style_dict = row_style_data.get(tgt_cell.column)
            if style_dict:
                tgt_cell.font = style_dict["font"]
                tgt_cell.border = style_dict["border"]
                tgt_cell.fill = style_dict["fill"]
                tgt_cell.number_format = style_dict["number_format"]
                tgt_cell.protection = style_dict["protection"]
                tgt_cell.alignment = style_dict["alignment"]