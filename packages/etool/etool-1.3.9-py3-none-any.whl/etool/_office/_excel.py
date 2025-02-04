import copy
import openpyxl
from openpyxl.utils import get_column_letter

class ManagerExcel:
    @staticmethod
    def excel_format(path, save_path):
        """
        复制excel文件，并保留格式
        :param path: 源文件路径
        :param save_path: 保存路径
        :return: 保存路径
        """
        wb = openpyxl.load_workbook(path)
        wb_new = openpyxl.Workbook()

        sheetnames = wb.sheetnames
        for sheetname in sheetnames:

            sheet = wb[sheetname]
            sheet2 = wb_new.create_sheet(sheetname)

            # 复制tab颜色
            sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

            # 开始处理合并单元格形式为“(,)，替换掉(,)' 找到合并单元格
            wm = list(sheet.merged_cells)
            if len(wm) > 0:
                for i in range(0, len(wm)):
                    cell2 = str(wm[i]).replace('(,)', '')
                    sheet2.merge_cells(cell2)

            # 遍历后，先写入数据
            for i, row in enumerate(sheet.iter_rows()):
                sheet2.row_dimensions[i+1].height = sheet.row_dimensions[i+1].height
                for j, cell in enumerate(row):
                    sheet2.column_dimensions[get_column_letter(
                        j+1)].width = sheet.column_dimensions[get_column_letter(j+1)].width
                    sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                    # 接着逐一设置单元格格式
                    source_cell = sheet.cell(i+1, j+1)
                    target_cell = sheet2.cell(i+1, j+1)
                    target_cell.fill = copy.copy(source_cell.fill)

                    # 默认样式是 Normal，如果是默认样式，返回False，不触发if，反之则进行复制
                    if source_cell.has_style: 

                        # 该StyleableObject实现将样式存储在单个列表中_style，并且单元格上的样式属性实际上是该数组的 getter 和 setter，所以你可以使用下方的写法，克隆样式更快
                        target_cell._style = copy.copy(source_cell._style)

                        # 复制字号
                        target_cell.font = copy.copy(source_cell.font)

                        # 复制边框
                        target_cell.border = copy.copy(source_cell.border)

                        # 复制填充样式
                        target_cell.fill = copy.copy(source_cell.fill)

                        # 复制字体样式
                        target_cell.number_format = copy.copy(
                            source_cell.number_format)

                        # 复制样式保护
                        target_cell.protection = copy.copy(source_cell.protection)

                        # 复制对齐样式
                        target_cell.alignment = copy.copy(source_cell.alignment)

        if 'Sheet' in wb_new.sheetnames:
            del wb_new['Sheet']
        wb_new.save(save_path)

        wb.close()
        wb_new.close()
        return save_path
