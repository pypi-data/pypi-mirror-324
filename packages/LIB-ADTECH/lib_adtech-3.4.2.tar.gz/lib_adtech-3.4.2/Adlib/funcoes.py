import os
from pathlib import Path
import sys
import time
import uuid
import shutil
import requests
import datetime
import subprocess
from PIL import Image
from Adlib.api import *
from urllib.parse import parse_qs, urlparse
from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.actions.action_builder import ActionBuilder


def setupDriver(
    webdrivePath: str = r"C:\Users\yan.fontes\Documents\webdriver\chromedriver-win32\chromedriver.exe", 
    numTabs: int = 1, 
    options: list[str] = [],
    experimentalOptions: list[str] = []
) -> Chrome:
    """
    Configura e inicializa uma instância do navegador Google Chrome utilizando o Selenium WebDriver.

    Args:
        webdrivePath (str): Caminho para o executável do ChromeDriver.
        numTabs (int): Número de guias a serem abertas no navegador.
        options (list[str]): Lista de argumentos adicionais para o ChromeOptions.

    Returns:
        Chrome: Uma instância configurada do WebDriver do Chrome pronta para ser usada em automações de teste ou navegação.
    """
    chrome_service = ChromeService(executable_path=webdrivePath)
    chrome_service.creation_flags = subprocess.CREATE_NO_WINDOW
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("log-level=3")

    for option in options:
        chrome_options.add_argument(option)

    driver = Chrome(service=chrome_service, options=chrome_options)
    driver.maximize_window()

    for _ in range(numTabs - 1):
        driver.execute_script("window.open('');")

    driver.switch_to.window(driver.window_handles[0])

    return driver


def getCredenciais(id: int) -> tuple[str, str] | tuple[None, None]:
    """
    Recupera as credenciais (login e senha) de uma API com base no ID fornecido.

    Esta função faz uma requisição `GET` para uma API REST usando o ID fornecido e tenta recuperar as credenciais de login e senha. Se a requisição for bem-sucedida (status code 200) e os dados estiverem presentes, ela retorna uma tupla contendo o login e a senha. Caso contrário, retorna uma tupla com `None` nos dois valores.

    Args:
        id (int): O ID utilizado para buscar as credenciais na API.

    Returns:
        tuple[str, str] | tuple[None, None]: 
            - Uma tupla contendo `login` e `senha` se a requisição for bem-sucedida e os dados estiverem presentes.
            - Uma tupla `(None, None)` se a requisição falhar ou os dados não estiverem disponíveis.
    """
    url = f"http://172.16.10.6:8080/credenciais/{id}"
    try:
        resposta = requests.get(url)
        if resposta.status_code == 200:
            dados = resposta.json()
            login = dados.get('login')
            senha = dados.get('senha')
            return login, senha
        return None, None
    except Exception as e:
        print(e)
        print("Não foi possível acessar a API")


def instalarPacote(pacote: str):
    """
    Instala uma biblioteca do python
    Arguments:
        pacote: nome do pacote disponível no PyPI
    """
    subprocess.check_call([sys.executable, "-m", "pip", "install", pacote])
 
 
def getNumeroSolicitacao(virtaus: Chrome):
    time.sleep(5)

    urlAtual = virtaus.current_url
 
    parsed_url = urlparse(urlAtual)
    query_params = parse_qs(parsed_url.query)
 
    if 'app_ecm_workflowview_processInstanceId' in query_params:
        return query_params['app_ecm_workflowview_processInstanceId'][0]
    return None
 

def aguardarAlert(driver: Chrome) -> str:
    try:
        alert = WebDriverWait(driver, 10).until(EC.alert_is_present())
        alert_text = alert.text
        try:
            alert.accept()
        except:
            alert.dismiss()
        return alert_text
    except:
        return ""


def selectOption(driver: webdriver.Chrome, selectXpath: str, value: str):
    proximaAtividade = Select(esperarElemento(driver, selectXpath))
    proximaAtividade.select_by_value(value)


def importarLibs():
    """Importa e instala, caso necessário, os pacotes e bibliotecas necessárias"""
    import time
    try:
        import pandas as pd
    except ImportError:
        instalarPacote("pandas")
        import pandas as pd
 
    try:
        from typing import Text
    except ImportError:
        instalarPacote("typing-extensions")
        from typing import Text
 
    try:
        from bs4 import element, BeautifulSoup, BeautifulStoneSoup
        from bs4.element import ProcessingInstruction
 
    except ImportError:
        instalarPacote("beautifulsoup4")
        from bs4 import element, BeautifulSoup, BeautifulStoneSoup
        from bs4.element import ProcessingInstruction
 
    try:
        from numpy import SHIFT_DIVIDEBYZERO, False_, exp
    except ImportError:
        instalarPacote("numpy")
        from numpy import SHIFT_DIVIDEBYZERO, False_, exp
 
    try:
        from pandas.io import html
    except ImportError:
        instalarPacote("pandas")
        from pandas.io import html
 
    try:
        from selenium import webdriver
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.action_chains import ActionChains
        from selenium.webdriver.chrome.options import Options
        
        
    except ImportError:
        instalarPacote("selenium")
        from selenium import webdriver
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.action_chains import ActionChains
        from selenium.webdriver.chrome.options import Options
        
        
 
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        instalarPacote("openpyxl")
        from openpyxl import Workbook, load_workbook
 
    try:
        import os
    except ImportError:
        instalarPacote("os")
        import os
 
    try:
        import tkinter as tk
        from tkinter import messagebox
    except ImportError:
        instalarPacote("tk")
        import tkinter as tk
        from tkinter import messagebox
 
    try:
        from discord_webhook import DiscordWebhook, DiscordEmbed
    except ImportError:
        instalarPacote("discord-webhook")
        from discord_webhook import DiscordWebhook, DiscordEmbed
 
    try:
        from datetime import datetime
    except ImportError:
        instalarPacote("datetime")
        from datetime import datetime
 
    bibliotecas = {
        'os.path': os.path,
        'pd': pd,
        'Text': Text,
        'element': element,
        'ProcessingInstruction': ProcessingInstruction,
        'SHIFT_DIVIDEBYZERO': SHIFT_DIVIDEBYZERO,
        'False_': False_,
        'exp': exp,
        'html': html,
        'webdriver': webdriver,
        'Keys': Keys,
        'ActionChains': ActionChains,
        'time': time,
        'DataFrame': pd.DataFrame,
        'read_csv': pd.read_csv,
        'Options': Options,
        'BeautifulSoup': BeautifulSoup,
        'BeautifulStoneSoup': BeautifulStoneSoup,
        'Workbook': Workbook,
        'load_workbook': load_workbook,
        'datetime': datetime,
        'tk': tk,
        'messagebox': messagebox,
        'DiscordWebhook': DiscordWebhook,
        'DiscordEmbed': DiscordEmbed,
        'WebDriverWait': WebDriverWait,
    }
 
    return bibliotecas


def esperarElemento(driver: Chrome, xpath: str, tempo_espera=10):
    """
    Aguarda o elemento ser renderizado
    Arguments:
        driver: driver do site
        xpath: XPath do elemento
        tempo_espera: Tempo máximo de espera, em segundos
    Returns:
        Elemento
    """
    try:
        return WebDriverWait(driver, tempo_espera).until(EC.visibility_of_element_located(('xpath', xpath)))
    except:
        print(f"Elemento não encontrado: {xpath}")
 
 
def esperarElementos(driver: Chrome, xpath: str, tempo_espera=10) -> list[WebElement]:
    """
    Aguarda todos os elementos serem renderizados.
    Arguments:
        driver: driver do site
        xpath: XPath dos elementos
        tempo_espera: Tempo máximo de espera, em segundos
    Returns:
        Lista de elementos
    """
    try:
        return WebDriverWait(driver, tempo_espera).until(EC.presence_of_all_elements_located(('xpath', xpath)))
    except:
        return []
 

def clickarElemento(driver: Chrome, xpath: str, time_wait=10):
    """
    Retorna o elemento do Xpath de entrada
    Args:
        driver: driver do site
        xpath: XPath do elemento
    Returns:
        Elemento
    """
    try:
        return WebDriverWait(driver, time_wait).until(EC.element_to_be_clickable(('xpath', xpath)))
    except:
        print(f"Elemento não encontrado: {xpath}")


def clickElement(driver: Chrome, xpath: str, tempoEspera: int = 20):
    """
        Aguarda o elemento entrar em estado clicável e executa um clique
    """
    driver.execute_script("arguments[0].click();", WebDriverWait(driver, tempoEspera).until(EC.element_to_be_clickable((By.XPATH, xpath))))


def mensagemTelegram(token: str, chat_id: int, mensagem: str):
    """
    Envia uma mensagem pela API do Telegram
    Arguments:
        token: token do bot do Telegram
        chat_id: id do chat
        mensagem: mensagem a ser enviada
    Returns:
        JSON com a resposta da requisição
    """
    mensagem_formatada = f'https://api.telegram.org/bot{token}/sendMessage?chat_id={chat_id}&parse_mode=HTML&text={mensagem}'
    resposta = requests.get(mensagem_formatada)
    return resposta.json()
 
 
def importar_bibliotecas():
    import pandas as pd
    from typing import Text
    from bs4 import element
    from bs4.element import ProcessingInstruction
    from numpy import SHIFT_DIVIDEBYZERO, False_, exp
    from pandas.io import html
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
    import time
    import pandas as pd
    from pandas.core.frame import DataFrame
    from pandas.io.parsers import read_csv
    from selenium import webdriver
    from selenium.webdriver.chrome.webdriver import WebDriver
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.support import expected_conditions
    from selenium.webdriver.support.wait import WebDriverWait
    from selenium.webdriver.chrome.options import Options
    from bs4 import BeautifulSoup, BeautifulStoneSoup
    from selenium import webdriver
    from openpyxl import Workbook, load_workbook
    import os
    from datetime import datetime
    from selenium.webdriver.chrome.service import Service
    from bs4 import BeautifulSoup, BeautifulStoneSoup
    from selenium import webdriver
    from openpyxl import Workbook, load_workbook
    from datetime import datetime
    import tkinter as tk
    from tkinter import messagebox
    from discord_webhook import DiscordWebhook, DiscordEmbed
    
    
    from selenium.webdriver.common.keys import Keys
 
    bibliotecas = {
        'os.path': os.path,
        'pd': pd,
        'Text': Text,
        'element': element,
        'ProcessingInstruction': ProcessingInstruction,
        'SHIFT_DIVIDEBYZERO': SHIFT_DIVIDEBYZERO,
        'False_': False_,
        'exp': exp,
        'html': html,
        'webdriver': webdriver,
        'Keys': Keys,
        'ActionChains': ActionChains,
        'time': time,
        'DataFrame': DataFrame,
        'read_csv': read_csv,
        'Options': Options,
        'BeautifulSoup': BeautifulSoup,
        'BeautifulStoneSoup': BeautifulStoneSoup,
        'Workbook': Workbook,
        'load_workbook': load_workbook,
        'datetime': datetime,
        'Service': Service,
        'tk': tk,
        'messagebox': messagebox,
        'DiscordWebhook': DiscordWebhook,
        'DiscordEmbed': DiscordEmbed,
        'WebDriverWait': WebDriverWait,
    }
 
    return bibliotecas
 
       
def esperar_elemento(driver: Chrome, xpath: str, tempo_espera=10):
    
    return WebDriverWait(driver, tempo_espera).until(EC.visibility_of_element_located(('xpath', xpath)))


def aguardarDownload(usuarioWindows: str, substringNomeArquivo: str) -> str:
    """
    Aguarda o download de um arquivo contendo uma substring específica no nome.

    Args:
        usuarioWindows (str): Nome do usuário do Windows.
        substringNomeArquivo (str): Substring que o arquivo baixado deve conter no nome.

    Returns:
        str: Caminho completo do arquivo baixado, se encontrado.
    """
    
    pastaDownloads = os.path.join(r"C:\Users", usuarioWindows, "Downloads")

    if not os.path.exists(pastaDownloads):
        raise FileNotFoundError(f"A pasta de downloads não foi encontrada: {pastaDownloads}")

    arquivosAnteriores = set(os.listdir(pastaDownloads))

    while True:
        arquivosAtuais = set(os.listdir(pastaDownloads))
        novosArquivos = arquivosAtuais - arquivosAnteriores

        for novoArquivo in novosArquivos:
            if substringNomeArquivo in novoArquivo:
                caminhoArquivo = os.path.join(pastaDownloads, novoArquivo)

                if not novoArquivo.endswith(".crdownload"):
                    return caminhoArquivo

        arquivosAnteriores = arquivosAtuais
        time.sleep(1)


def enviarCaptcha(token: str, chatId: int, imagePath: str | Path, enumBanco: EnumBanco, enumProcesso: EnumProcesso, tempoEspera: int = 30) -> str:
    """
    Envia uma imagem do captcha para um chat do Telegram e retorna uma resposta no intervalo de tempo.

    Args:
        chat_id (int): ID do chat do Telegram.
        imagePath (str | Path): Caminho da imagem do captcha.
    """

    token = "6647704641:AAFmqTsrRlUdrM7EzCwlAktnttm8k1kJ6Q8"
    chatId = "-4095757991"
    

    baseUrl = f'https://api.telegram.org/bot{token}'

    formatName = lambda x: (" ".join(c for c in x.split('_'))).upper()

    with open(imagePath, 'rb') as imageFile:
        parametros = {
            "chat_id": chatId,
            "caption": f"Realizar Captcha\n {formatName(enumBanco.name)} {formatName(enumProcesso.name)}"
        }

        files = {
            "photo": imageFile
        }

        resp = requests.post(f"{baseUrl}/sendPhoto", data=parametros, files=files).json()
        messageId = resp["result"]["message_id"]

    time.sleep(tempoEspera)
    response = requests.get(f"{baseUrl}/getUpdates")
    updates = response.json()
    
    for update in reversed(updates.get("result", [])):
        if "message" in update and update["message"]["reply_to_message"]["message_id"] == messageId:
            return update["message"]["text"]
    
    return "123456"


def saveCaptchaImage(driver: Chrome, imgElement: WebElement, enumBanco: EnumBanco, enumProcesso: EnumProcesso):


    location = imgElement.location
    size = imgElement.size

    imgName = f"Token{enumBanco.name}{enumProcesso.name}.png"
    driver.save_screenshot(imgName)

    x = location['x']
    y = location['y']
    width = size['width']
    height = size['height']
    im = Image.open(imgName)
    im = im.crop((int(x), int(y), int(x + width), int(y + height)))
    
    os.remove(imgName)

    imgFolderPath = os.getcwd()

    imgPath = os.path.join(imgFolderPath, imgName)
    imageId = str(uuid.uuid4())
    imgPath = os.path.join(imgFolderPath, f'imgName_{imageId}.png')

    im.save(imgPath)

    return imgPath


def aguardarRespostaCatpcha():
    pass


def clickCoordenada(driver: Chrome, x: int, y: int) -> None:
    """
    Clica em uma coordenada específica na tela.
    Args:
        driver: driver do site
        x: coordenada x
        y: coordenada y
    """

    action = ActionBuilder(driver)
    action.pointer_action.move_to_location(x, y)
    action.pointer_action.click()
    action.perform()
    

def importarPastaMonitoramento(filePathList: list[str], diretorioBase: str, data: datetime.datetime = None):
    try:
        if not data:
            hoje = data
        else:
            hoje = datetime.datetime.now()
        pastaAno = str(hoje.year)
        pastaMes = f"{hoje.month:02d} - {[hoje.month]}"
        pastaDia = f"{hoje.day:02d}"
    
        caminho = os.path.join(diretorioBase, pastaAno, pastaMes, pastaDia)

        os.makedirs(caminho, exist_ok=True)

        for filePath in filePathList:
            
            nomeArquivo = os.path.basename(filePath)
            destino = os.path.join(caminho, nomeArquivo)

            shutil.copy(filePath, destino)
            os.remove(filePath)

        return caminho
    
    except Exception as e:
        print(e)

if __name__=="__main__":

    imgPath = r"C:\Users\dannilo.costa\Pictures\Screenshots\Captura de tela 2025-01-08 171740.png"
    print(enviarCaptcha(None, None, imgPath, EnumBanco.BMG, EnumProcesso.CRIACAO))