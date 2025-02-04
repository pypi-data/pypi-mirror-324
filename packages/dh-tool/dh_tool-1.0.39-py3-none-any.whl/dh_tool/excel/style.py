from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from .utils import map_column_names_to_letters

COLOR_MAP = {
    "black": "000000",
    "white": "FFFFFF",
    "red": "FF0000",
    "green": "00FF00",
    "blue": "0000FF",
    "yellow": "FFFF00",
    "cyan": "00FFFF",
    "magenta": "FF00FF",
    "gray": "808080",
    "orange": "FFA500",
    "purple": "800080",
    "pink": "FFC0CB",
    "brown": "A52A2A",
    "gold": "FFD700",
    "silver": "C0C0C0",
}


class Style:
    @staticmethod
    def apply_auto_wrap(target):
        for cell in (cell for row in target for cell in row):
            if cell.value:
                cell.alignment = Alignment(wrap_text=True)

    @staticmethod
    def set_column_width(worksheet, width_map):
        for col, width in width_map.items():
            worksheet.column_dimensions[col].width = width

    @staticmethod
    def set_font(target, font_name="Arial", font_size=12, bold=False, italic=False):
        font = Font(name=font_name, size=font_size, bold=bold, italic=italic)
        for cell in (cell for row in target for cell in row):
            cell.font = font

    @staticmethod
    def apply_border(target, border_style="thin"):
        border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style),
        )
        for cell in (cell for row in target for cell in row):
            cell.border = border

    @staticmethod
    def apply_color(target, color):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in (cell for row in target for cell in row):
            if cell.value:
                cell.fill = fill

    @staticmethod
    def freeze_first_row(worksheet):
        worksheet.freeze_panes = worksheet["A2"]

    @staticmethod
    def auto_adjust_column_widths(worksheet):
        for col in worksheet.columns:
            max_length = max(
                (len(str(cell.value)) for cell in col if cell.value), default=0
            )
            worksheet.column_dimensions[col[0].column_letter].width = max_length + 2
