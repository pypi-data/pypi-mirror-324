from openpyxl.utils import get_column_letter


def map_column_names_to_letters(worksheet, width_map):
    """
    DataFrame의 컬럼 이름 또는 엑셀 열 문자(A, B, C)를 자동 매핑하여 열 너비 설정
    """
    # 엑셀 시트의 헤더 가져오기
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    # 최종 매핑 결과 저장
    col_letter_map = {}

    for key, width in width_map.items():
        if key in headers:
            # ✅ 컬럼 이름을 엑셀 열 문자로 변환
            col_idx = headers.index(key) + 1
            col_letter = get_column_letter(col_idx)
            col_letter_map[col_letter] = width
        else:
            print(f"컬럼 '{key}'을 찾을 수 없습니다.")

    return col_letter_map


def get_column_indices_by_condition(worksheet, condition):
    """
    조건을 만족하는 열의 인덱스를 반환
    - condition: 각 열의 데이터 리스트를 받아 True/False를 반환하는 함수
    """
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    indices = []

    for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
        col_data = [cell.value for cell in col]
        if condition(col_data):
            indices.append(idx)

    return indices


def find_columns_with_nulls(worksheet):
    """
    결측치가 있는 컬럼 반환
    """
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    null_columns = []

    for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
        col_data = [cell.value for cell in col]
        if any(pd.isnull(value) for value in col_data):
            null_columns.append(headers[idx - 1])

    return null_columns


def find_columns_by_type(worksheet, data_type):
    """
    특정 데이터 타입(int, str 등)을 가진 열 찾기
    """
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    type_columns = []

    for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
        col_data = [cell.value for cell in col]
        if all(isinstance(value, data_type) or pd.isnull(value) for value in col_data):
            type_columns.append(headers[idx - 1])

    return type_columns
