from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


class Style:
    @staticmethod
    def apply_auto_wrap(worksheet):
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

    @staticmethod
    def set_column_width(worksheet, width_map):
        for col, width in width_map.items():
            worksheet.column_dimensions[col].width = width

    @staticmethod
    def apply_color(worksheet, color):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.fill = fill

    @staticmethod
    def freeze_first_row(worksheet):
        worksheet.freeze_panes = worksheet["A2"]

    @staticmethod
    def set_font(worksheet, font_name="Arial", font_size=12, bold=False, italic=False):
        font = Font(name=font_name, size=font_size, bold=bold, italic=italic)
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.font = font

    @staticmethod
    def apply_border(worksheet, border_style="thin"):
        border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style),
        )
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = border
