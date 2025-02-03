import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from netmiko import ConnectHandler

def style_switch_sheet(sheet, switch_name):
    """
    Applies styling to each per-switch sheet:
      - Inserts a merged row at the top with the switch_name centered.
      - Column widths (approx):
         A -> ~10.4
         B -> ~11.9
         C -> ~18.6
         D -> ~27.9
         E -> ~32.9
         F -> ~17
      - Row height = 16 for all rows.
      - Font: Bahnschrift, size=12, center alignment.
      - Highlight the Info cell if value="up" (yellow) or contains "admin" (orange).
      - Highlight the NAC cell if value="closed" (greenish) or "monitored" (reddish).
    """
    # 1) Insert merged title row at the top
    sheet.insert_rows(1)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = switch_name
    title_cell.font = Font(name="Bahnschrift", size=12, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # White background

    # 2) Set column widths
    sheet.column_dimensions['A'].width = 10.4  # Port
    sheet.column_dimensions['B'].width = 11.9  # Vlan-ID
    sheet.column_dimensions['C'].width = 18.6  # Info
    sheet.column_dimensions['D'].width = 27.9  # Device
    sheet.column_dimensions['E'].width = 32.9  # MAC
    sheet.column_dimensions['F'].width = 17.0  # NAC

    # 3) Set row heights and base styling
    for row_idx in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 16
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Bahnschrift", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Make row 2 (the header) bold
            if row_idx == 2:
                cell.font = Font(name="Bahnschrift", size=12, bold=True)

    # 4) Highlight Info and NAC cells with conditions
    for row_idx in range(3, sheet.max_row + 1):
        # Info column = C (3)
        info_cell = sheet.cell(row=row_idx, column=3)
        if isinstance(info_cell.value, str):
            lower_info = info_cell.value.lower()
            if lower_info == "up":
                info_cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")  # Yellow
            elif "admin" in lower_info:
                info_cell.fill = PatternFill(fill_type="solid", fgColor="FF5D37")  # Orange

        # NAC column = F (6)
        nac_cell = sheet.cell(row=row_idx, column=6)
        if isinstance(nac_cell.value, str):
            val = nac_cell.value.lower()
            if val == "closed":
                nac_cell.fill = PatternFill(fill_type="solid", fgColor="C6EFCE")  # Greenish
                nac_cell.font = Font(name="Bahnschrift", size=12, color="006100")
            elif val == "monitored":
                nac_cell.fill = PatternFill(fill_type="solid", fgColor="FFC7CE")  # Reddish
                nac_cell.font = Font(name="Bahnschrift", size=12, color="9C0006")


def style_summary_sheet(sheet):
    """
    Applies styling to the "Switches" summary sheet:
      - Font: Bahnschrift, size=12, row height=22
      - Headers (row 1): bold, white text on #305496
      - All data rows: fill with white
      - Column widths adjusted:
         A: 24
         B: 16
         C: 26.5
         D: 12.7
         E: 29
         F: 30
         G: 39
         H: 10
         I: 42
         J: 42
    """
    # 1) Set approximate column widths
    sheet.column_dimensions['A'].width = 24.0  # Switch
    sheet.column_dimensions['B'].width = 16.0  # IP
    sheet.column_dimensions['C'].width = 26.5  # type
    sheet.column_dimensions['D'].width = 12.7  # Version
    sheet.column_dimensions['E'].width = 29.0  # Model
    sheet.column_dimensions['F'].width = 30.0  # Serial
    sheet.column_dimensions['G'].width = 39.0  # Module MAC
    sheet.column_dimensions['H'].width = 10.0  # Stack
    sheet.column_dimensions['I'].width = 42.0  # Location
    sheet.column_dimensions['J'].width = 42.0  # Uptime

    # 2) Row height, alignment, base font
    for row_idx in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 22
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Bahnschrift", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 3) Header row => bold, white text on #305496
    header_fill = PatternFill(fill_type="solid", fgColor="305496")
    white_font = Font(name="Bahnschrift", size=12, bold=True, color="FFFFFF")

    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = white_font

    # 4) Data rows => white background
    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")


def get_switch_info(workbook, host, username, password, device_type="cisco_ios"):
    """
    Connects to a switch, gathers data, creates a sheet with columns:
      [Port, Vlan-ID, Info, Device, MAC, NAC]
    Returns a summary dict with switch details for the summary sheet.

    NAC logic:
      - For each interface, run 'show run interface {iface}'
      - If 'access-session close' -> NAC='closed'
      - If 'dot1x pae authenticator' -> NAC='monitored'
      - Else -> NAC='Off'
    """
    device = {
        "device_type": device_type,
        "host": host,
        "username": username,
        "password": password,
    }

    switch_data = {}

    try:
        connection = ConnectHandler(**device)
        print(f"Connected to {host}")

        # 1) Get Hostname
        run_output = connection.send_command("show run")
        hostname = host  # fallback
        for line in run_output.splitlines():
            if line.startswith("hostname "):
                _, hname = line.split(None, 1)
                hostname = hname.strip()
                break

        # 2) SNMP Location
        location_cmd = connection.send_command("show run | sec snmp-server location")
        match_loc = re.search(r"snmp-server location\s+(.*)", location_cmd)
        location_value = match_loc.group(1).strip() if match_loc else "N/A"

        # 3) show mod => gather lines (stack info, etc.)
        mod_out = connection.send_command("show mod")
        model_list, serial_list, mac_list, ver_list = [], [], [], []

        for line in mod_out.splitlines():
            # Skip header / separator lines
            if (
                ("Switch" in line and "Ports" in line and "Model" in line) or
                ("-----" in line) or
                ("Mod" in line and "Ports" in line)
            ):
                continue

            match_mod = re.match(
                r"^\s*(\d+)\s+(\d+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)",
                line
            )
            if match_mod:
                model_list.append(match_mod.group(3))
                serial_list.append(match_mod.group(4))
                mac_list.append(match_mod.group(5))
                ver_list.append(match_mod.group(7))

        if not model_list:
            model_list  = ["N/A"]
            serial_list = ["N/A"]
            mac_list    = ["N/A"]
            ver_list    = ["N/A"]

        stack_count = len(serial_list)
        stack_value = str(stack_count) if stack_count > 1 else "No"
        version_str = ver_list[0] if ver_list else "N/A"

        switch_type_str   = ", ".join(model_list)
        serial_number_str = ", ".join(serial_list)
        module_mac_str    = ", ".join(mac_list)

        # 4) Uptime from show version
        ver_out = connection.send_command("show version")
        uptime_value = "N/A"
        for line in ver_out.splitlines():
            if "uptime is" in line:
                # e.g. "C9300-24P uptime is 52 days, 3 hours, 1 minute"
                uptime_value = line.split("uptime is", 1)[1].strip()
                break

        # 5) IP address
        ip_address = host

        # 6) VLAN + trunk info from show interface switchport
        sw_out = connection.send_command("show interface switchport")
        vlan_map = {}
        trunk_native_vlan = {}
        current_interface = None

        for line in sw_out.splitlines():
            if line.startswith("Name:"):
                current_interface = line.split("Name:")[1].strip()

            elif "Access Mode VLAN" in line:
                m = re.search(r"Access Mode VLAN:\s+(\d+)", line)
                if current_interface and m:
                    acc_vlan = int(m.group(1))
                    # if VLAN=1, treat as trunk in some envs
                    vlan_map[current_interface] = (
                        "Trunk" if acc_vlan == 1 else str(acc_vlan)
                    )
            elif "Trunking Native Mode VLAN" in line:
                m = re.search(r"Trunking Native Mode VLAN:\s+(\d+)", line)
                if current_interface and m:
                    native_vlan = m.group(1)
                    trunk_native_vlan[current_interface] = native_vlan
                    if current_interface not in vlan_map:
                        vlan_map[current_interface] = "Trunk"

        # 7) show interfaces description
        desc_out = connection.send_command("show interfaces description")
        interface_list = []
        for line in desc_out.splitlines():
            if (not line.strip()) or line.startswith("Interface"):
                continue

            parts = re.split(r"\s{2,}", line.strip(), maxsplit=3)
            if len(parts) >= 3:
                iface = parts[0]
                status = parts[1]
                protocol = parts[2]
                desc = parts[3] if len(parts) == 4 else ""

                if iface.startswith("Gi"):
                    interface_list.append({
                        "Interface": iface,
                        "Status":    status,
                        "Device":    desc,
                        "VLAN":      vlan_map.get(iface, "N/A")
                    })

        # 8) MAC address info
        mac_map_by_int_vlan = {}

        # a) dynamic MACs
        mac_out_dynamic = connection.send_command("show mac address-table dynamic")
        for line in mac_out_dynamic.splitlines():
            if re.match(r"^\s*\d+\s+[a-fA-F0-9]{4}\.[a-fA-F0-9]{4}\.[a-fA-F0-9]{4}", line):
                parts = line.split()
                if len(parts) >= 4:
                    vlan_id, mac, _type, port = parts[:4]
                    mac_map_by_int_vlan[(port, vlan_id)] = mac

        # b) static MACs
        mac_out_static = connection.send_command("show mac address-table static")
        for line in mac_out_static.splitlines():
            if re.match(r"^\s*\d+\s+[a-fA-F0-9]{4}\.[a-fA-F0-9]{4}\.[a-fA-F0-9]{4}", line):
                parts = line.split()
                if len(parts) >= 4:
                    vlan_id, mac, _type, port = parts[:4]
                    if (port, vlan_id) not in mac_map_by_int_vlan:
                        mac_map_by_int_vlan[(port, vlan_id)] = mac

        # 8.5) NAC status
        nac_map = {}
        for entry in interface_list:
            iface = entry["Interface"]
            nac_map[iface] = "Off"
            int_config = connection.send_command(f"show run interface {iface}")
            if "access-session close" in int_config:
                nac_map[iface] = "closed"
            elif "dot1x pae authenticator" in int_config:
                nac_map[iface] = "monitored"

        # 9) Create the per-switch sheet
        ws = workbook.create_sheet(title=hostname)
        ws.append(["Port", "Vlan-ID", "Info", "Device", "MAC", "NAC"])

        for entry in interface_list:
            iface   = entry["Interface"]
            status  = entry["Status"]
            device_ = entry["Device"]
            vlan    = entry["VLAN"]

            mac_address = "N/A"
            if vlan.isdigit():
                mac_address = mac_map_by_int_vlan.get((iface, vlan), "N/A")
            elif vlan == "Trunk":
                if iface in trunk_native_vlan:
                    native_vlan = trunk_native_vlan[iface]
                    mac_address = mac_map_by_int_vlan.get((iface, native_vlan), "N/A")

            nac_status = nac_map.get(iface, "Off")
            ws.append([iface, vlan, status, device_, mac_address, nac_status])

        # 10) Apply styling
        style_switch_sheet(ws, hostname)

        # 11) Prepare summary data
        switch_data[hostname] = {
            "IP Address":    ip_address,
            "Switch Type":   switch_type_str,
            "Version":       version_str,
            "Model":         switch_type_str,
            "Serial Number": serial_number_str,
            "Module MAC":    module_mac_str,
            "Stack":         stack_value,
            "Location":      location_value,
            "Uptime":        uptime_value
        }

        connection.disconnect()
        print(f"Data for switch '{hostname}' added successfully.\n")

    except Exception as e:
        print(f"Failed to connect to {host}: {e}")

    return switch_data


def create_summary_sheet(workbook, switch_data):
    """
    Creates a sheet named "Switches" with columns:
      Switch, IP Address, type, Version, Model, Serial Number,
      Module MAC, Stack, Location, Uptime
    Then applies the summary sheet styling.
    """
    summary_sheet = workbook.create_sheet(title="Switches")
    summary_sheet.append([
        "Switch",
        "IP Address",
        "type",
        "Version",
        "Model",
        "Serial Number",
        "Module MAC",
        "Stack",
        "Location",
        "Uptime"
    ])

    for hostname, data in switch_data.items():
        summary_sheet.append([
            hostname,
            data["IP Address"],
            data["Switch Type"],
            data["Version"],
            data["Model"],
            data["Serial Number"],
            data["Module MAC"],
            data["Stack"],
            data["Location"],
            data["Uptime"]
        ])

    style_summary_sheet(summary_sheet)
